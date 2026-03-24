"""
CRE Enterprise Suite - PDF Export & Excel Import Module
Handles document generation and data ingestion
"""
import io
import json
import uuid
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any

# PDF Generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor, black, white
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

# Excel Processing
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# Colors
DARK_BG = HexColor('#0F1419') if HAS_PDF else None
BLUE = HexColor('#3B82F6') if HAS_PDF else None
GREEN = HexColor('#10B981') if HAS_PDF else None
AMBER = HexColor('#F59E0B') if HAS_PDF else None
RED = HexColor('#EF4444') if HAS_PDF else None
LIGHT_GRAY = HexColor('#F3F4F6') if HAS_PDF else None
DARK_GRAY = HexColor('#374151') if HAS_PDF else None
WHITE = HexColor('#FFFFFF') if HAS_PDF else None

# ============ PDF GENERATOR ============
class PDFGenerator:
    
    @staticmethod
    def generate_investment_memo(property_data: Dict, rent_roll: Dict, valuation: Dict) -> bytes:
        """Generate professional investment memorandum PDF"""
        if not HAS_PDF:
            return b"PDF generation not available. Install reportlab: pip install reportlab"
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=HexColor('#1F2937'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=HexColor('#6B7280'),
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=HexColor('#1F2937'),
            spaceBefore=16,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            textColor=HexColor('#374151'),
            spaceAfter=6
        )
        
        # ---- HEADER ----
        story.append(Paragraph("INVESTMENT MEMORANDUM", title_style))
        story.append(Paragraph(f"Prepared: {datetime.now().strftime('%B %d, %Y')}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#3B82F6')))
        story.append(Spacer(1, 0.2*inch))
        
        # ---- PROPERTY OVERVIEW ----
        story.append(Paragraph("PROPERTY OVERVIEW", heading_style))
        
        prop_data = [
            ['Property Name', property_data.get('name', 'N/A')],
            ['Asset Type', property_data.get('asset_type', 'N/A').title()],
            ['Address', property_data.get('address', 'N/A')],
            ['Purchase Price', f"${float(property_data.get('purchase_price', 0)):,.2f}"],
            ['Purchase Date', property_data.get('purchase_date', 'N/A')],
        ]
        
        prop_table = Table(prop_data, colWidths=[2*inch, 4.5*inch])
        prop_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#EFF6FF')),
            ('TEXTCOLOR', (0, 0), (0, -1), HexColor('#1D4ED8')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [WHITE, LIGHT_GRAY]),
        ]))
        story.append(prop_table)
        story.append(Spacer(1, 0.2*inch))
        
        # ---- FINANCIAL METRICS ----
        story.append(Paragraph("FINANCIAL METRICS", heading_style))
        
        irr = valuation.get('irr', 0)
        coc = valuation.get('coc', 0)
        dscr = valuation.get('dscr', 0)
        cap_rate = valuation.get('cap_rate', 0)
        noi = valuation.get('noi', 0)
        
        metrics_data = [
            ['Metric', 'Value', 'Benchmark', 'Status'],
            ['IRR (Internal Rate of Return)', f"{irr}%", '> 12%', '✓ GOOD' if irr > 12 else '⚠ REVIEW'],
            ['CoC (Cash-on-Cash Return)', f"{coc}%", '> 8%', '✓ GOOD' if coc > 8 else '⚠ REVIEW'],
            ['DSCR (Debt Service Coverage)', f"{dscr}", '> 1.25', '✓ GOOD' if dscr > 1.25 else '⚠ REVIEW'],
            ['Cap Rate', f"{cap_rate}%", '5-8%', '✓ GOOD' if 5 <= cap_rate <= 8 else '⚠ REVIEW'],
            ['NOI (Net Operating Income)', f"${noi:,.2f}", 'N/A', ''],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.6*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#3B82F6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 0.2*inch))
        
        # ---- RENT ROLL SUMMARY ----
        if rent_roll and rent_roll.get('entries'):
            story.append(Paragraph("RENT ROLL SUMMARY", heading_style))
            
            summary_data = [
                ['Total Units', str(rent_roll.get('total_units', 0))],
                ['Occupied Units', str(rent_roll.get('occupied_units', 0))],
                ['Occupancy Rate', f"{rent_roll.get('occupancy_rate', 0)}%"],
                ['Total Monthly Rent', f"${rent_roll.get('total_monthly_rent', 0):,.2f}"],
                ['Annual Rent', f"${rent_roll.get('total_monthly_rent', 0) * 12:,.2f}"],
            ]
            
            summary_table = Table(summary_data, colWidths=[2*inch, 4.5*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), HexColor('#ECFDF5')),
                ('TEXTCOLOR', (0, 0), (0, -1), HexColor('#065F46')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [WHITE, LIGHT_GRAY]),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.15*inch))
            
            # Unit detail table
            story.append(Paragraph("Unit Details:", body_style))
            unit_data = [['Unit', 'Tenant', 'Monthly Rent', 'Lease End', 'Status']]
            for entry in rent_roll.get('entries', [])[:20]:  # Max 20 units
                unit_data.append([
                    entry.get('unit_number', ''),
                    entry.get('tenant_name', 'Vacant'),
                    f"${entry.get('monthly_rent', 0):,.0f}",
                    entry.get('lease_end', 'N/A'),
                    entry.get('status', 'vacant').upper()
                ])
            
            unit_table = Table(unit_data, colWidths=[0.8*inch, 1.8*inch, 1.2*inch, 1.2*inch, 1.5*inch])
            unit_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('PADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ]))
            story.append(unit_table)
            story.append(Spacer(1, 0.2*inch))
        
        # ---- INVESTMENT THESIS ----
        story.append(Paragraph("INVESTMENT THESIS", heading_style))
        
        thesis_points = [
            f"• Strong NOI of ${noi:,.2f} provides stable cash flow foundation",
            f"• IRR of {irr}% {'exceeds' if irr > 12 else 'approaches'} target return threshold of 12%",
            f"• DSCR of {dscr} {'comfortably covers' if dscr > 1.25 else 'meets minimum'} debt service requirements",
            f"• Cap rate of {cap_rate}% is {'competitive' if 5 <= cap_rate <= 8 else 'notable'} in current market",
            f"• Asset type: {property_data.get('asset_type', 'N/A').title()} offers {'stable' if property_data.get('asset_type') == 'multifamily' else 'strong'} demand fundamentals",
        ]
        
        for point in thesis_points:
            story.append(Paragraph(point, body_style))
        
        story.append(Spacer(1, 0.2*inch))
        
        # ---- RISK FACTORS ----
        story.append(Paragraph("RISK FACTORS", heading_style))
        
        risks = [
            "• Market risk: Changes in local real estate market conditions",
            "• Interest rate risk: Rising rates may impact refinancing",
            "• Vacancy risk: Tenant turnover may affect cash flow",
            "• Regulatory risk: Changes in zoning or rent control laws",
            "• Capital expenditure risk: Unexpected maintenance costs",
        ]
        
        for risk in risks:
            story.append(Paragraph(risk, body_style))
        
        story.append(Spacer(1, 0.2*inch))
        
        # ---- FOOTER ----
        story.append(HRFlowable(width="100%", thickness=1, color=HexColor('#E5E7EB')))
        story.append(Spacer(1, 0.1*inch))
        
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=HexColor('#9CA3AF'),
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            f"CONFIDENTIAL - CRE Enterprise Suite | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | For authorized use only",
            footer_style
        ))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.read()
    
    @staticmethod
    def generate_rent_roll_pdf(property_name: str, rent_roll: Dict) -> bytes:
        """Generate rent roll PDF report"""
        if not HAS_PDF:
            return b"PDF generation not available"
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()
        story = []
        
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, textColor=HexColor('#1F2937'), fontName='Helvetica-Bold')
        heading_style = ParagraphStyle('Heading', parent=styles['Heading1'], fontSize=14, textColor=HexColor('#1F2937'), fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=8)
        
        story.append(Paragraph(f"RENT ROLL REPORT", title_style))
        story.append(Paragraph(f"{property_name} | {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#10B981')))
        story.append(Spacer(1, 0.2*inch))
        
        # Summary
        story.append(Paragraph("SUMMARY", heading_style))
        summary_data = [
            ['Total Units', str(rent_roll.get('total_units', 0)), 'Occupied', str(rent_roll.get('occupied_units', 0))],
            ['Occupancy Rate', f"{rent_roll.get('occupancy_rate', 0)}%", 'Monthly Rent', f"${rent_roll.get('total_monthly_rent', 0):,.2f}"],
        ]
        summary_table = Table(summary_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#ECFDF5')),
            ('BACKGROUND', (2, 0), (2, -1), HexColor('#ECFDF5')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Unit details
        story.append(Paragraph("UNIT DETAILS", heading_style))
        unit_data = [['Unit #', 'Tenant', 'Monthly Rent', 'Lease Start', 'Lease End', 'Status']]
        for entry in rent_roll.get('entries', []):
            unit_data.append([
                entry.get('unit_number', ''),
                entry.get('tenant_name', 'Vacant'),
                f"${entry.get('monthly_rent', 0):,.0f}",
                entry.get('lease_start', 'N/A'),
                entry.get('lease_end', 'N/A'),
                entry.get('status', 'vacant').upper()
            ])
        
        unit_table = Table(unit_data, colWidths=[0.8*inch, 1.8*inch, 1.1*inch, 1.1*inch, 1.1*inch, 0.8*inch])
        unit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
        ]))
        story.append(unit_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.read()
    
    @staticmethod
    def generate_valuation_report(property_data: Dict, valuation: Dict) -> bytes:
        """Generate valuation report PDF"""
        if not HAS_PDF:
            return b"PDF generation not available"
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=0.75*inch, leftMargin=0.75*inch, topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()
        story = []
        
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, textColor=HexColor('#1F2937'), fontName='Helvetica-Bold')
        heading_style = ParagraphStyle('Heading', parent=styles['Heading1'], fontSize=14, textColor=HexColor('#1F2937'), fontName='Helvetica-Bold', spaceBefore=12, spaceAfter=8)
        
        story.append(Paragraph("VALUATION REPORT", title_style))
        story.append(Paragraph(f"{property_data.get('name', 'Property')} | {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
        story.append(HRFlowable(width="100%", thickness=2, color=HexColor('#3B82F6')))
        story.append(Spacer(1, 0.2*inch))
        
        # Key metrics
        story.append(Paragraph("KEY METRICS", heading_style))
        
        metrics = [
            ['IRR', f"{valuation.get('irr', 0)}%", 'Internal Rate of Return'],
            ['CoC', f"{valuation.get('coc', 0)}%", 'Cash-on-Cash Return'],
            ['DSCR', f"{valuation.get('dscr', 0)}", 'Debt Service Coverage Ratio'],
            ['Cap Rate', f"{valuation.get('cap_rate', 0)}%", 'Capitalization Rate'],
            ['NOI', f"${valuation.get('noi', 0):,.2f}", 'Net Operating Income'],
        ]
        
        metrics_table = Table(metrics, colWidths=[1.2*inch, 1.5*inch, 3.8*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#EFF6FF')),
            ('TEXTCOLOR', (0, 0), (0, -1), HexColor('#1D4ED8')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#E5E7EB')),
            ('ROWBACKGROUNDS', (0, 0), (-1, -1), [WHITE, LIGHT_GRAY]),
        ]))
        story.append(metrics_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.read()


# ============ EXCEL PROCESSOR ============
class ExcelProcessor:
    
    @staticmethod
    def parse_rent_roll(file_content: bytes) -> Dict[str, Any]:
        """Parse rent roll from Excel file"""
        if not HAS_EXCEL:
            return {"error": "Excel processing not available. Install openpyxl: pip install openpyxl"}
        
        try:
            buffer = io.BytesIO(file_content)
            df = pd.read_excel(buffer)
            
            # Normalize column names
            df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
            
            # Map common column names
            column_map = {
                'unit': 'unit_number',
                'unit_no': 'unit_number',
                'unit_#': 'unit_number',
                'tenant': 'tenant_name',
                'name': 'tenant_name',
                'rent': 'monthly_rent',
                'monthly_rent': 'monthly_rent',
                'lease_start': 'lease_start',
                'start_date': 'lease_start',
                'lease_end': 'lease_end',
                'end_date': 'lease_end',
                'status': 'status',
            }
            
            df = df.rename(columns=column_map)
            
            entries = []
            for _, row in df.iterrows():
                entry = {
                    'unit_number': str(row.get('unit_number', '')),
                    'tenant_name': str(row.get('tenant_name', 'Vacant')),
                    'monthly_rent': float(row.get('monthly_rent', 0) or 0),
                    'lease_start': str(row.get('lease_start', '')),
                    'lease_end': str(row.get('lease_end', '')),
                    'status': str(row.get('status', 'occupied')).lower(),
                }
                entries.append(entry)
            
            total_rent = sum(e['monthly_rent'] for e in entries)
            occupied = len([e for e in entries if e['status'] == 'occupied'])
            
            return {
                'entries': entries,
                'total_units': len(entries),
                'occupied_units': occupied,
                'total_monthly_rent': total_rent,
                'occupancy_rate': round((occupied / len(entries) * 100) if entries else 0, 2),
                'status': 'success'
            }
        except Exception as e:
            return {'error': str(e), 'status': 'error'}
    
    @staticmethod
    def export_rent_roll(property_name: str, rent_roll: Dict) -> bytes:
        """Export rent roll to Excel"""
        if not HAS_EXCEL:
            return b"Excel export not available"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rent Roll"
        
        # Styles
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1F2937")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"RENT ROLL - {property_name.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Summary
        ws['A4'] = "SUMMARY"
        ws['A4'].font = Font(bold=True, size=12)
        ws['A5'] = "Total Units:"
        ws['B5'] = rent_roll.get('total_units', 0)
        ws['C5'] = "Occupied:"
        ws['D5'] = rent_roll.get('occupied_units', 0)
        ws['A6'] = "Occupancy Rate:"
        ws['B6'] = f"{rent_roll.get('occupancy_rate', 0)}%"
        ws['C6'] = "Monthly Rent:"
        ws['D6'] = f"${rent_roll.get('total_monthly_rent', 0):,.2f}"
        
        # Headers
        headers = ['Unit #', 'Tenant Name', 'Monthly Rent', 'Lease Start', 'Lease End', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        for row_idx, entry in enumerate(rent_roll.get('entries', []), 9):
            ws.cell(row=row_idx, column=1, value=entry.get('unit_number', ''))
            ws.cell(row=row_idx, column=2, value=entry.get('tenant_name', 'Vacant'))
            ws.cell(row=row_idx, column=3, value=entry.get('monthly_rent', 0))
            ws.cell(row=row_idx, column=4, value=entry.get('lease_start', ''))
            ws.cell(row=row_idx, column=5, value=entry.get('lease_end', ''))
            ws.cell(row=row_idx, column=6, value=entry.get('status', '').upper())
            
            if row_idx % 2 == 0:
                for col in range(1, 7):
                    ws.cell(row=row_idx, column=col).fill = alt_fill
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    
    @staticmethod
    def export_valuation_report(property_name: str, valuation: Dict) -> bytes:
        """Export valuation report to Excel"""
        if not HAS_EXCEL:
            return b"Excel export not available"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Valuation"
        
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color="1F2937")
        
        ws.merge_cells('A1:D1')
        ws['A1'] = f"VALUATION REPORT - {property_name.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        headers = ['Metric', 'Value', 'Benchmark', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        irr = valuation.get('irr', 0)
        coc = valuation.get('coc', 0)
        dscr = valuation.get('dscr', 0)
        cap_rate = valuation.get('cap_rate', 0)
        noi = valuation.get('noi', 0)
        
        metrics = [
            ['IRR', f"{irr}%", '> 12%', 'GOOD' if irr > 12 else 'REVIEW'],
            ['Cash-on-Cash', f"{coc}%", '> 8%', 'GOOD' if coc > 8 else 'REVIEW'],
            ['DSCR', f"{dscr}", '> 1.25', 'GOOD' if dscr > 1.25 else 'REVIEW'],
            ['Cap Rate', f"{cap_rate}%", '5-8%', 'GOOD' if 5 <= cap_rate <= 8 else 'REVIEW'],
            ['NOI', f"${noi:,.2f}", 'N/A', ''],
        ]
        
        green_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        red_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
        
        for row_idx, metric in enumerate(metrics, 5):
            for col, val in enumerate(metric, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if col == 4:
                    cell.fill = green_fill if val == 'GOOD' else red_fill
                    cell.font = Font(bold=True, color="065F46" if val == 'GOOD' else "991B1B")
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    
    @staticmethod
    def parse_property_data(file_content: bytes) -> Dict[str, Any]:
        """Parse property data from Excel"""
        if not HAS_EXCEL:
            return {"error": "Excel processing not available"}
        
        try:
            buffer = io.BytesIO(file_content)
            df = pd.read_excel(buffer)
            df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
            
            properties = []
            for _, row in df.iterrows():
                prop = {
                    'name': str(row.get('name', row.get('property_name', ''))),
                    'asset_type': str(row.get('asset_type', row.get('type', 'multifamily'))).lower(),
                    'address': str(row.get('address', '')),
                    'purchase_price': float(row.get('purchase_price', row.get('price', 0)) or 0),
                    'purchase_date': str(row.get('purchase_date', row.get('date', ''))),
                }
                properties.append(prop)
            
            return {'properties': properties, 'total': len(properties), 'status': 'success'}
        except Exception as e:
            return {'error': str(e), 'status': 'error'}
